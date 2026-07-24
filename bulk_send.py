"""
FLASH Smaranotsavam 2026 - Bulk WhatsApp Template Sender (Python version)
Sends the approved "ambulance_inauguration_" template with video header
to every number in the contact list.

Run with: python bulk_send.py
Requires: pip install requests openpyxl
"""

import requests
import time
import csv
import os
from openpyxl import load_workbook

# ====== CONFIG - fill these in ======
PHONE_NUMBER_ID = "1211131482079268"
ACCESS_TOKEN = "PASTE_YOUR_ACCESS_TOKEN_HERE"
TEMPLATE_NAME = "ambulance_inauguration_"  # confirm exact name in WhatsApp Manager
MEDIA_ID = "2497321997439570"
EXCEL_FILE = "Unique_Mobile_Numbers_1.xlsx"
DELAY_SECONDS = 1.2
LOG_FILE = "send_log.csv"
TEST_BATCH_ONLY = False   # set True first to test on a small batch
TEST_BATCH_SIZE = 20
# ======================================

URL = f"https://graph.facebook.com/v20.0/{PHONE_NUMBER_ID}/messages"


def send_template(number):
    body = {
        "messaging_product": "whatsapp",
        "to": number,
        "type": "template",
        "template": {
            "name": TEMPLATE_NAME,
            "language": {"code": "te"},
            "components": [
                {
                    "type": "header",
                    "parameters": [{"type": "video", "video": {"id": MEDIA_ID}}]
                }
            ]
        }
    }
    headers = {
        "Authorization": f"Bearer {ACCESS_TOKEN}",
        "Content-Type": "application/json"
    }
    resp = requests.post(URL, json=body, headers=headers)
    return resp.status_code, resp.json()


def load_numbers():
    wb = load_workbook(EXCEL_FILE, read_only=True)
    ws = wb.active
    numbers = []
    header = None
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = row
            continue
        # "Mobile Number" is expected in column B (index 1)
        num = row[1]
        if num:
            numbers.append("91" + str(num))
    return numbers


def load_already_sent():
    sent = set()
    if os.path.exists(LOG_FILE):
        with open(LOG_FILE, newline="") as f:
            reader = csv.reader(f)
            next(reader, None)  # skip header
            for row in reader:
                if row and row[1] == "SUCCESS":
                    sent.add(row[0])
    else:
        with open(LOG_FILE, "w", newline="") as f:
            csv.writer(f).writerow(["number", "status", "message_id_or_error"])
    return sent


def main():
    numbers = load_numbers()

    if TEST_BATCH_ONLY:
        numbers = numbers[:TEST_BATCH_SIZE]
        print(f"TEST MODE: only sending to first {len(numbers)} numbers.")

    print(f"Loaded {len(numbers)} numbers. Starting send...")

    already_sent = load_already_sent()
    if already_sent:
        print(f"Resuming: {len(already_sent)} already sent, skipping those.")

    sent_count = 0
    fail_count = 0

    with open(LOG_FILE, "a", newline="") as f:
        writer = csv.writer(f)

        for number in numbers:
            if number in already_sent:
                continue

            try:
                status, data = send_template(number)

                if status == 200 and "messages" in data:
                    writer.writerow([number, "SUCCESS", data["messages"][0]["id"]])
                    sent_count += 1
                else:
                    err = data.get("error", data)
                    writer.writerow([number, "FAILED", str(err)])
                    fail_count += 1

                    if isinstance(err, dict) and err.get("code") == 130429:
                        print("Rate limit hit. Pausing 60 seconds...")
                        time.sleep(60)

            except Exception as e:
                writer.writerow([number, "ERROR", str(e)])
                fail_count += 1

            f.flush()

            total_done = sent_count + fail_count
            if total_done % 100 == 0:
                remaining = len(numbers) - total_done
                print(f"Progress: {sent_count} sent, {fail_count} failed, {remaining} remaining")

            time.sleep(DELAY_SECONDS)

    print(f"\nDone. Total sent: {sent_count}, Total failed: {fail_count}")
    print(f"Full results in {LOG_FILE}")


if __name__ == "__main__":
    main()
